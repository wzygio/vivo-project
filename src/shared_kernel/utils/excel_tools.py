# src/vivo_project/utils/utils.py
import pandas as pd
import logging
import os
import sys
import tempfile
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Optional
import streamlit as st  # [新增] 引入 streamlit

from src.shared_kernel.config import ConfigLoader


def read_workbook_sheet(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """读取共享工作簿中的指定 sheet。

    文件不存在或 sheet 不存在时返回空 DataFrame（与“文件缺失”语义一致）；
    企业加密等 openpyxl 无法读取的工作簿自动回退到 Excel COM 透明解密读取，
    COM 也失败时抛出异常交由调用方决定（告警或包装为业务异常）。

    Args:
        xlsx_path: 共享工作簿路径
        sheet_name: 目标 sheet 名（通常为产品号或 <产品号>_<原sheet名>）

    Returns:
        该 sheet 的数据；文件/sheet 缺失时为空 DataFrame
    """
    if not xlsx_path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(xlsx_path, sheet_name=sheet_name)
    except ValueError:
        # Worksheet named 'X' not found —— sheet 缺失视为无数据
        return pd.DataFrame()
    except Exception as openpyxl_error:
        logging.warning(
            "[excel_tools] 标准读取 %s [%s] 失败，尝试 Excel COM: %s",
            xlsx_path.name, sheet_name, openpyxl_error,
        )
        return _read_encrypted_xlsx_via_com(xlsx_path, sheet_name)


def _read_all_sheets_via_com(xlsx_path: Path) -> dict[str, pd.DataFrame]:
    """通过 Excel COM 读出工作簿的全部 sheets（用于加密工作簿的整体重写回退）。"""
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        raise ImportError("win32com 未安装，无法读取加密工作簿。")

    com_initialized = False
    try:
        pythoncom.CoInitialize()
        com_initialized = True
    except Exception:
        pass

    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()), ReadOnly=True)
        sheets: dict[str, pd.DataFrame] = {}
        for ws in wb.Worksheets:
            data = ws.UsedRange.Value
            if data is None or len(data) < 1:
                sheets[ws.Name] = pd.DataFrame()
                continue
            headers = list(data[0])
            rows = [list(row) for row in data[1:]]
            sheets[ws.Name] = pd.DataFrame(rows, columns=headers)
        return sheets
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        wb = None
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        excel = None
        if com_initialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


@dataclass(frozen=True)
class WorkbookWriteResult:
    """工作簿写入结果契约（PRD §5.10）。

    - written: 是否真正完成了正式文件的原子替换；
    - path: 目标工作簿路径；
    - updated_sheets: 本次提交中更新（替换/新建）的 sheet 名，按传入顺序；
    - error: 失败时的可操作错误信息，成功时为 None。
    """

    written: bool
    path: Path
    updated_sheets: tuple[str, ...]
    error: str | None = None


# 同一工作簿的写入在进程内互斥，避免并发写产生部分更新
_workbook_write_lock = threading.Lock()


def replace_workbook_sheets(
    xlsx_path: Path,
    sheets: Mapping[str, pd.DataFrame],
) -> WorkbookWriteResult:
    """在一次工作簿提交中原子替换（或新建）多个 sheet，保留其他 sheet。

    事务语义：所有目标 sheet 要么全部生效，要么全部不生效——
    先在同目录临时文件中完成保存并用 openpyxl 回读验证
    （目标 sheet 存在且行数符合预期），再 os.replace 原子替换正式文件；
    任一环节失败均返回 written=False，正式文件保持不变。

    - 企业加密等 openpyxl 无法打开的文件回退为：COM 读出全部 sheets，
      合并更新后整体重写为明文 xlsx（logger.warning 明确记录）；
    - 文件被占用（PermissionError，如 Excel 打开中）返回 written=False，
      error 含“请关闭 Excel 后重试”提示，不抛错；
    - 正式文件不存在时按同样流程新建。

    Args:
        xlsx_path: 共享工作簿路径
        sheets: {目标 sheet 名: 写入数据}，顺序即 updated_sheets 顺序

    Returns:
        WorkbookWriteResult，调用方必须显式检查 written
    """
    with _workbook_write_lock:
        return _replace_workbook_sheets_locked(xlsx_path, sheets)


def _replace_workbook_sheets_locked(
    xlsx_path: Path,
    sheets: Mapping[str, pd.DataFrame],
) -> WorkbookWriteResult:
    """replace_workbook_sheets 的实际实现（调用方需已持有 _workbook_write_lock）。"""
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    target_sheets = tuple(sheets.keys())
    if not target_sheets:
        return WorkbookWriteResult(
            written=False, path=xlsx_path, updated_sheets=(), error="未提供需要写入的 sheet"
        )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    # 同目录临时文件：保证 os.replace 为同卷原子替换
    tmp_handle = tempfile.NamedTemporaryFile(
        prefix=f".{xlsx_path.stem}-", suffix=".tmp.xlsx", dir=xlsx_path.parent, delete=False
    )
    tmp_path = Path(tmp_handle.name)
    tmp_handle.close()

    try:
        wb = None
        if xlsx_path.exists():
            try:
                wb = openpyxl.load_workbook(xlsx_path)
            except Exception:
                wb = None

        if wb is None and xlsx_path.exists():
            # openpyxl 无法打开（企业加密等）：COM 读出全部 sheets 后整体重写为明文
            logging.warning(
                "[excel_tools] openpyxl 无法打开 %s，企业加密工作簿整体重写为明文 xlsx。",
                xlsx_path.name,
            )
            try:
                all_sheets = _read_all_sheets_via_com(xlsx_path)
            except Exception as exc:
                return WorkbookWriteResult(
                    written=False,
                    path=xlsx_path,
                    updated_sheets=(),
                    error=f"COM 读取加密工作簿失败: {exc}",
                )
            for name, df in sheets.items():
                all_sheets[name] = df
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                for name, sheet_df in all_sheets.items():
                    sheet_df.to_excel(writer, index=False, sheet_name=name)
        else:
            if wb is None:
                # 正式文件不存在：新建工作簿
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
            for name, df in sheets.items():
                if name in wb.sheetnames:
                    del wb[name]
                ws = wb.create_sheet(name)
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
            wb.save(tmp_path)

        # 回读临时文件验证：目标 sheet 存在且行数（含表头）符合预期
        verify_error = _verify_temp_workbook(tmp_path, sheets)
        if verify_error is not None:
            return WorkbookWriteResult(
                written=False, path=xlsx_path, updated_sheets=(), error=verify_error
            )

        os.replace(tmp_path, xlsx_path)
        return WorkbookWriteResult(
            written=True, path=xlsx_path, updated_sheets=target_sheets
        )
    except PermissionError as exc:
        return WorkbookWriteResult(
            written=False,
            path=xlsx_path,
            updated_sheets=(),
            error=f"工作簿被占用（可能被 Excel 打开），请关闭 Excel 后重试: {exc}",
        )
    except Exception as exc:
        return WorkbookWriteResult(
            written=False,
            path=xlsx_path,
            updated_sheets=(),
            error=f"写入工作簿失败: {exc}",
        )
    finally:
        # 无论成败都清理临时文件；已原子替换成功时 tmp_path 已不存在
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass


def _verify_temp_workbook(
    tmp_path: Path, sheets: Mapping[str, pd.DataFrame]
) -> str | None:
    """回读临时工作簿验证目标 sheet 存在且行数（含表头）符合预期；失败返回错误信息。"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
    except Exception as exc:
        return f"临时工作簿回读验证失败: {exc}"
    try:
        for name, df in sheets.items():
            if name not in wb.sheetnames:
                return f"临时工作簿验证失败: 目标 sheet [{name}] 缺失"
            expected_rows = len(df) + 1  # 含表头行
            actual_rows = wb[name].max_row or 0
            if actual_rows != expected_rows:
                return (
                    f"临时工作簿验证失败: sheet [{name}] 行数 {actual_rows}，"
                    f"预期 {expected_rows}"
                )
    finally:
        wb.close()
    return None


def replace_workbook_sheet(xlsx_path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    """在共享工作簿中替换（或新建）指定 sheet，保留其他 sheet 的内容与格式。

    [兼容包装] 内部委托 replace_workbook_sheets；新代码请直接使用
    replace_workbook_sheets 并显式检查 WorkbookWriteResult.written。

    兼容语义：失败（如文件被 Excel 占用）时仅记录告警，不抛错。

    Args:
        xlsx_path: 共享工作簿路径（不存在则新建单 sheet 工作簿）
        sheet_name: 目标 sheet 名
        df: 写入的数据
    """
    result = replace_workbook_sheets(xlsx_path, {sheet_name: df})
    if not result.written:
        logging.warning(
            "[excel_tools] 写入工作簿 %s [%s] 失败: %s",
            xlsx_path, sheet_name, result.error,
        )


def _read_encrypted_xlsx_via_com(xlsx_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    [COM fallback] 通过 Windows Excel.Application COM 接口读取加密/受保护 xlsx 文件。
    
    企业加密软件会锁定 xlsx 文件使 openpyxl/xlrd 无法直接读取，
    但 Windows 本地 Excel 应用程序能够透明解密并打开这些文件。
    
    Args:
        xlsx_path: 加密的 xlsx 文件路径
        sheet_name: 可选的 sheet 名称（默认读取第一个 sheet）
    
    Returns:
        包含解密后数据的 DataFrame
    
    Raises:
        ImportError: win32com 未安装
        Exception: Excel 打开或读取失败
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        raise ImportError(
            "win32com 未安装。请执行 `pip install pywin32` 后重试，"
            "或确保在解密环境中运行 xlsx_to_csv。"
        )
    
    com_initialized = False
    try:
        pythoncom.CoInitialize()
        com_initialized = True
    except Exception:
        pass

    excel = None
    wb = None
    ws = None
    used_range = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()), ReadOnly=True)
        if sheet_name:
            ws = wb.Worksheets(sheet_name)
        else:
            ws = wb.Worksheets(1)
        
        used_range = ws.UsedRange
        data = used_range.Value

        if data is None or len(data) < 1:
            return pd.DataFrame()

        # COM 返回 tuple of tuples，首行为表头
        headers = list(data[0])
        rows = [list(row) for row in data[1:]]
        df = pd.DataFrame(rows, columns=headers)

        logging.info(f"[xlsx_to_csv] 使用 COM (Excel.Application) 读取加密文件 {xlsx_path.name} 成功, shape={df.shape}")
        return df

    except Exception as e:
        logging.error(f"[xlsx_to_csv] COM 读取加密文件失败: {e}")
        raise
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        used_range = None
        ws = None
        wb = None
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        excel = None
        if com_initialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def xlsx_to_csv(
    xlsx_path: Path,
    csv_dir: Path,
    sheet_name: Optional[str] = None,
    encoding: str = "utf-8-sig",
    **read_excel_kwargs,
) -> Path:
    """
    [通用工具] 将 xlsx 文件转换为 csv 格式。

    用途：在企业加密环境中，xlsx 可能被加密导致 openpyxl 无法读取，
          提前在解密环境中将其转为 csv，供生产代码作为 fallback 读取。
          
          [增强] 当 openpyxl/xlrd 均无法读取时（如企业加密软件锁定），
          自动尝试 Windows COM (Excel.Application) 接口进行透明解密读取。

    Args:
        xlsx_path: 源 xlsx 文件路径
        csv_dir:   输出 csv 目录
        sheet_name: 指定 sheet 名（单 sheet 文件可留空）
        encoding:  csv 编码，默认 utf-8-sig（Excel 兼容）
        **read_excel_kwargs: 透传给 pd.read_excel 的额外参数

    Returns:
        生成的 csv 文件路径
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"源文件不存在: {xlsx_path}")

    csv_dir.mkdir(parents=True, exist_ok=True)
    csv_path = csv_dir / xlsx_path.with_suffix(".csv").name

    # 读取 xlsx：优先用 openpyxl，失败时尝试 xlrd
    df = pd.DataFrame()
    engines = ["openpyxl", "xlrd"]
    last_err = None
    for engine in engines:
        try:
            kwargs = dict(read_excel_kwargs)
            if sheet_name is not None:
                kwargs["sheet_name"] = sheet_name
            df = pd.read_excel(xlsx_path, engine=engine, **kwargs)
            logging.info(f"[xlsx_to_csv] 使用引擎 {engine} 读取 {xlsx_path.name} 成功, shape={df.shape}")
            break
        except Exception as e:
            last_err = e
            logging.warning(f"[xlsx_to_csv] 引擎 {engine} 读取失败: {e}")
            continue

    # [增强] 引擎全部失败时，尝试 COM fallback（应对企业加密软件锁定）
    if df.empty and last_err is not None:
        try:
            df = _read_encrypted_xlsx_via_com(xlsx_path, sheet_name)
            logging.info(f"[xlsx_to_csv] COM fallback 读取成功, shape={df.shape}")
        except Exception as com_err:
            logging.error(f"[xlsx_to_csv] COM fallback 也失败: {com_err}")
            raise last_err from com_err

    if df.empty:
        raise ValueError(f"无法从 {xlsx_path.name} 读取到任何数据")

    # 写入 csv（不写入 index，保留原始表头）
    df.to_csv(csv_path, index=False, encoding=encoding)
    logging.info(f"[xlsx_to_csv] 成功导出: {csv_path}")
    return csv_path


def batch_export_spc_rules_to_csv(
    resource_dir: Optional[Path] = None,
    csv_subdir: str = "xlsx_to_csv",
) -> list[Path]:
    """
    [批量导出] 将 SPC 相关的加密 xlsx 规则文件批量导出为 csv。
    目标文件：spc_outlier_filters.xlsx

    应在能透明解密企业加密文件的环境中执行一次，
    之后生产代码即可通过 csv fallback 读取规则。

    Returns:
        生成的 csv 文件路径列表
    """
    if resource_dir is None:
        resource_dir = ConfigLoader.get_project_root() / "resources"

    target_files = ["spc_outlier_filters.xlsx"]
    csv_dir = resource_dir / csv_subdir
    exported: list[Path] = []

    for fname in target_files:
        xlsx_path = resource_dir / fname
        if not xlsx_path.exists():
            logging.warning(f"[batch_export_spc_rules_to_csv] 跳过不存在的文件: {xlsx_path}")
            continue
        try:
            csv_path = xlsx_to_csv(xlsx_path, csv_dir)
            exported.append(csv_path)
        except Exception as e:
            logging.error(f"[batch_export_spc_rules_to_csv] 导出 {fname} 失败: {e}")

    return exported


def save_dict_to_excel(data_dict: dict, output_dir: Path, filename: str):
    """
    [通用工具] 将包含 DataFrame 的字典保存到 Excel。
    (此函数保持原样，无需修改)
    """
    if not isinstance(data_dict, dict) or not data_dict:
        logging.error(f"[调试] 无法保存 {filename}：输入不是有效的字典或字典为空！")
        return

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = output_dir / filename

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            logging.info(f"[调试探针] 正在将数据写入 {file_path}...")
            saved_sheets_count = 0

            for key, value in data_dict.items():
                if isinstance(value, pd.DataFrame) and not value.empty:
                    sheet_name = str(key)
                    clean_sheet_name = sheet_name.replace(':', '_').replace('/', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
                    if len(clean_sheet_name) > 31: clean_sheet_name = clean_sheet_name[:31]

                    try:
                        value.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                        saved_sheets_count += 1
                        logging.debug(f"Saved top-level DataFrame '{key}' to sheet '{clean_sheet_name}'.")
                    except Exception as sheet_error:
                        logging.error(f"[调试] 写入 Sheet 页 '{clean_sheet_name}' (来自顶层键 '{key}') 时出错: {sheet_error}")

                elif key == 'code_level_details' and isinstance(value, dict):
                    logging.debug("Found 'code_level_details', iterating inner dictionary...")
                    for group_name, group_df in value.items():
                        if isinstance(group_df, pd.DataFrame) and not group_df.empty:
                            sheet_name = str(group_name)
                            clean_sheet_name = sheet_name.replace(':', '_').replace('/', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
                            if len(clean_sheet_name) > 31: clean_sheet_name = clean_sheet_name[:31]

                            try:
                                group_df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                                saved_sheets_count += 1
                                logging.debug(f"Saved inner DataFrame '{group_name}' to sheet '{clean_sheet_name}'.")
                            except Exception as sheet_error:
                                logging.error(f"[调试] 写入 Sheet 页 '{clean_sheet_name}' (来自 code_level_details['{group_name}']) 时出错: {sheet_error}")

            if saved_sheets_count > 0:
                logging.info(f"[调试探针] 成功将 {saved_sheets_count} 个 DataFrame 保存到: {file_path}")
            else:
                logging.warning(f"[调试] 未能在字典中找到有效的 DataFrame 以保存到 {filename}。")

    except Exception as e:
        logging.error(f"[调试] 保存调试 Excel 文件 '{filename}' 时发生错误: {e}", exc_info=True)

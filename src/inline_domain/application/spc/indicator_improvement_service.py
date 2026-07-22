# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import logging
import re
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Optional
from zipfile import BadZipFile

import pandas as pd

from src.inline_domain.core.spc.indicator_spec_comparison import (
    compare_specs,
    constraints_to_text,
    make_display_name,
    make_spec_identity,
    normalize_cell,
    parse_monitor_spec,
)

logger = logging.getLogger(__name__)

PRODUCT_ORDER: tuple[str, ...] = ("M678", "M626", "Z571")
EARLY_VERSION_PRODUCT_MAP: dict[str, str] = {
    "20251127": "M678",
    "20260205": "M626",
}
REQUIRED_COLUMNS: tuple[str, ...] = ("厂别", "科室", "监控因子", "描述", "监控规格")
IDENTITY_COLUMNS: tuple[str, ...] = (
    "厂别",
    "科室",
    "工艺（膜层）",
    "工序（站点）",
    "监控因子",
    "RS_CODE",
    "描述",
)
FILL_DOWN_COLUMNS: tuple[str, ...] = (
    "厂别",
    "科室",
    "工艺（膜层）",
    "工序（站点）",
    "失效风险",
    "监控因子",
    "监控方案",
    "抽检频率",
    "RS_CODE",
    "描述",
    "实例",
)


@dataclass(frozen=True)
class IndicatorSpecRecord:
    source_file: str
    version: str
    product: str
    sheet_name: str
    source_row: int
    identity_key: str
    display_name: str
    factory: str
    department: str
    process_layer: str
    station: str
    monitor_factor: str
    rs_code: str
    description: str
    raw_spec: str
    parsed_spec: str
    comparable: bool
    parse_reason: str


@dataclass(frozen=True)
class IndicatorAnalysisResult:
    output_dir: Path
    workbook_path: Path
    task1_image_path: Path
    task2_image_path: Path
    task1_summary: pd.DataFrame
    task2_summary: pd.DataFrame
    task1_details: pd.DataFrame
    task2_details: pd.DataFrame


class IndicatorImprovementService:
    def __init__(
        self,
        source_dir: Path,
        output_dir: Path,
        normalized_workbook_dir: Path,
        products: tuple[str, ...] = PRODUCT_ORDER,
        early_version_product_map: Optional[dict[str, str]] = None,
    ) -> None:
        self.source_dir = source_dir
        self.output_dir = output_dir
        self.products = products
        self.early_version_product_map = early_version_product_map or EARLY_VERSION_PRODUCT_MAP
        self.normalized_workbook_dir = normalized_workbook_dir

    def run(self) -> IndicatorAnalysisResult:
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            self.normalized_workbook_dir.mkdir(parents=True, exist_ok=True)
            source_files = self._discover_source_files()
            records = self._extract_records(source_files)
            extracted_specs_df = pd.DataFrame([asdict(record) for record in records])
            task1_summary_df, task1_details_df, task1_all_df = self._build_task1_tables(records)
            task2_summary_df, task2_details_df, task2_all_df = self._build_task2_tables(records)
            plot_config_df = self._build_plot_config(task1_summary_df, task2_summary_df)
            metadata_df = self._build_metadata(source_files, extracted_specs_df)

            self._write_csv_outputs(
                extracted_specs_df=extracted_specs_df,
                task1_summary_df=task1_summary_df,
                task1_details_df=task1_details_df,
                task1_all_df=task1_all_df,
                task2_summary_df=task2_summary_df,
                task2_details_df=task2_details_df,
                task2_all_df=task2_all_df,
                plot_config_df=plot_config_df,
            )
            workbook_path = self._write_excel_outputs(
                extracted_specs_df=extracted_specs_df,
                task1_summary_df=task1_summary_df,
                task1_details_df=task1_details_df,
                task1_all_df=task1_all_df,
                task2_summary_df=task2_summary_df,
                task2_details_df=task2_details_df,
                task2_all_df=task2_all_df,
                plot_config_df=plot_config_df,
                metadata_df=metadata_df,
            )
            task1_image_path = self._draw_task1_chart(plot_config_df)
            task2_image_path = self._draw_task2_chart(plot_config_df)
            self._write_metadata_json(source_files, extracted_specs_df, task1_summary_df, task2_summary_df)

            return IndicatorAnalysisResult(
                output_dir=self.output_dir,
                workbook_path=workbook_path,
                task1_image_path=task1_image_path,
                task2_image_path=task2_image_path,
                task1_summary=task1_summary_df,
                task2_summary=task2_summary_df,
                task1_details=task1_details_df,
                task2_details=task2_details_df,
            )
        except Exception as exc:
            logger.exception("[IndicatorImprovement] 任务执行失败: %s", exc)
            raise

    def _discover_source_files(self) -> list[Path]:
        try:
            files = [
                item
                for item in self.source_dir.glob("*.xlsx")
                if item.is_file() and not item.name.startswith("~$") and _extract_version(item.name)
            ]
            return sorted(files, key=lambda item: _extract_version(item.name) or "")
        except Exception as exc:
            logger.exception("[IndicatorImprovement] 扫描指标文件失败: %s", exc)
            raise

    def _extract_records(self, source_files: list[Path]) -> list[IndicatorSpecRecord]:
        records: list[IndicatorSpecRecord] = []
        for source_file in source_files:
            version = _extract_version(source_file.name)
            if version is None:
                continue
            sheets = self._read_workbook_sheets(source_file)
            self._write_normalized_workbook(source_file, version)
            targets = self._resolve_product_sheets(version, sheets)
            for product, sheet_name in targets.items():
                sheet_records = self._records_from_sheet(
                    rows=sheets[sheet_name],
                    source_file=source_file.name,
                    version=version,
                    product=product,
                    sheet_name=sheet_name,
                )
                records.extend(sheet_records)
        return records

    def _read_workbook_sheets(self, path: Path) -> dict[str, list[list[Any]]]:
        try:
            return _read_workbook_with_openpyxl(path)
        except Exception as openpyxl_error:
            logger.warning("[IndicatorImprovement] openpyxl 读取失败，切换 COM: %s", openpyxl_error)
            try:
                return _read_workbook_with_com(path)
            except Exception as com_error:
                logger.exception("[IndicatorImprovement] COM 读取失败: %s", com_error)
                raise openpyxl_error from com_error

    def _write_normalized_workbook(self, source_file: Path, version: str) -> None:
        output_path = self.normalized_workbook_dir / f"indicator_workbook_{version}.xlsx"
        try:
            if output_path.exists():
                output_path.unlink()
            try:
                _read_workbook_with_openpyxl(source_file)
                shutil.copy2(source_file, output_path)
            except Exception:
                _save_workbook_with_com(source_file, output_path)
        except Exception as exc:
            logger.warning("[IndicatorImprovement] 规范化工作簿输出失败，继续使用内存数据: %s", exc)

    def _resolve_product_sheets(self, version: str, sheets: dict[str, list[list[Any]]]) -> dict[str, str]:
        targets: dict[str, str] = {}
        sheet_names = list(sheets)
        for product in self.products:
            matched = next((name for name in sheet_names if product in name), None)
            if matched is not None:
                targets[product] = matched

        if targets:
            return targets

        mapped_product = self.early_version_product_map.get(version)
        generic_sheet = next((name for name in sheet_names if "指标拆解" in name), None)
        if mapped_product and generic_sheet:
            return {mapped_product: generic_sheet}
        return {}

    def _records_from_sheet(
        self,
        rows: list[list[Any]],
        source_file: str,
        version: str,
        product: str,
        sheet_name: str,
    ) -> list[IndicatorSpecRecord]:
        header_index = _find_header_index(rows)
        if header_index is None:
            return []

        header = [normalize_cell(value) for value in rows[header_index]]
        column_indexes = {name: index for index, name in enumerate(header) if name}
        if any(required not in column_indexes for required in REQUIRED_COLUMNS):
            return []

        last_values: dict[str, str] = {column: "" for column in FILL_DOWN_COLUMNS}
        occurrence_counts: dict[str, int] = {}
        records: list[IndicatorSpecRecord] = []
        for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values: dict[str, str] = {}
            for column in set(FILL_DOWN_COLUMNS) | set(REQUIRED_COLUMNS) | set(IDENTITY_COLUMNS):
                index = column_indexes.get(column)
                raw_value = row[index] if index is not None and index < len(row) else None
                value = normalize_cell(raw_value)
                if column in FILL_DOWN_COLUMNS:
                    if value:
                        last_values[column] = value
                    else:
                        value = last_values.get(column, "")
                values[column] = value

            raw_spec = normalize_cell(row[column_indexes["监控规格"]] if column_indexes["监控规格"] < len(row) else None)
            if raw_spec == "":
                continue
            if not any(values.get(column, "") for column in IDENTITY_COLUMNS):
                continue

            base_identity = "|".join(values.get(column, "NA") or "NA" for column in IDENTITY_COLUMNS)
            occurrence_counts[base_identity] = occurrence_counts.get(base_identity, 0) + 1
            occurrence = occurrence_counts[base_identity]
            parsed = parse_monitor_spec(raw_spec)
            identity_key = make_spec_identity(
                factory=values.get("厂别", ""),
                department=values.get("科室", ""),
                process_layer=values.get("工艺（膜层）", ""),
                station=values.get("工序（站点）", ""),
                monitor_factor=values.get("监控因子", ""),
                rs_code=values.get("RS_CODE", ""),
                description=values.get("描述", ""),
                occurrence=occurrence,
            )
            display_name = make_display_name(
                factory=values.get("厂别", ""),
                department=values.get("科室", ""),
                monitor_factor=values.get("监控因子", ""),
                description=values.get("描述", ""),
                process_layer=values.get("工艺（膜层）", ""),
                station=values.get("工序（站点）", ""),
                rs_code=values.get("RS_CODE", ""),
                occurrence=occurrence,
            )
            records.append(
                IndicatorSpecRecord(
                    source_file=source_file,
                    version=version,
                    product=product,
                    sheet_name=sheet_name,
                    source_row=offset,
                    identity_key=identity_key,
                    display_name=display_name,
                    factory=values.get("厂别", ""),
                    department=values.get("科室", ""),
                    process_layer=values.get("工艺（膜层）", ""),
                    station=values.get("工序（站点）", ""),
                    monitor_factor=values.get("监控因子", ""),
                    rs_code=values.get("RS_CODE", ""),
                    description=values.get("描述", ""),
                    raw_spec=raw_spec,
                    parsed_spec=constraints_to_text(parsed.constraints),
                    comparable=parsed.comparable,
                    parse_reason=parsed.reason,
                )
            )
        return records

    def _build_task1_tables(
        self,
        records: list[IndicatorSpecRecord],
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        comparisons: list[dict[str, Any]] = []
        details: list[dict[str, Any]] = []
        by_product_version = _index_records(records)

        for product in self.products:
            versions = sorted(version for prod, version in by_product_version if prod == product)
            for old_version, new_version in zip(versions, versions[1:]):
                old_records = by_product_version[(product, old_version)]
                new_records = by_product_version[(product, new_version)]
                result_rows = _compare_record_maps(
                    task="Task1",
                    comparison_label=f"{product}: {old_version} -> {new_version}",
                    old_label=old_version,
                    new_label=new_version,
                    old_records=old_records,
                    new_records=new_records,
                )
                comparisons.extend(result_rows)
                details.extend(row for row in result_rows if row["is_tightened"])

        all_df = pd.DataFrame(comparisons)
        details_df = pd.DataFrame(details)
        summary_df = _summarize_comparisons(all_df, group_columns=["product", "from_version", "to_version"])
        return summary_df, details_df, all_df

    def _build_task2_tables(
        self,
        records: list[IndicatorSpecRecord],
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        comparisons: list[dict[str, Any]] = []
        details: list[dict[str, Any]] = []
        by_product_version = _index_records(records)
        latest_version = max(version for _product, version in by_product_version)

        for old_product, new_product in zip(self.products, self.products[1:]):
            old_records = by_product_version.get((old_product, latest_version), {})
            new_records = by_product_version.get((new_product, latest_version), {})
            result_rows = _compare_record_maps(
                task="Task2",
                comparison_label=f"{new_product} vs {old_product} ({latest_version})",
                old_label=old_product,
                new_label=new_product,
                old_records=old_records,
                new_records=new_records,
            )
            for row in result_rows:
                row["version"] = latest_version
                row["base_product"] = old_product
                row["new_product"] = new_product
            comparisons.extend(result_rows)
            details.extend(row for row in result_rows if row["is_tightened"])

        all_df = pd.DataFrame(comparisons)
        details_df = pd.DataFrame(details)
        summary_df = _summarize_comparisons(all_df, group_columns=["version", "base_product", "new_product"])
        return summary_df, details_df, all_df

    def _build_plot_config(self, task1_summary_df: pd.DataFrame, task2_summary_df: pd.DataFrame) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        colors = ["#1E88E5", "#43A047", "#FB8C00", "#8E24AA", "#00ACC1", "#E53935"]
        for index, row in task1_summary_df.reset_index(drop=True).iterrows():
            label = f"{row.get('product', '')} {row.get('from_version', '')}->{row.get('to_version', '')}"
            rows.append(
                {
                    "chart_id": "task1_version_tightening",
                    "task": "Task1",
                    "series": str(row.get("product", "")),
                    "category": label,
                    "metric": "tightened_count",
                    "value": int(row.get("tightened_count", 0)),
                    "label": label,
                    "color": colors[index % len(colors)],
                    "sort_order": index + 1,
                    "include": True,
                    "title": "不同版本指标收严成果",
                    "subtitle": "同一产品相较上一版本的可比规格收严项数",
                }
            )
        for index, row in task2_summary_df.reset_index(drop=True).iterrows():
            label = f"{row.get('new_product', '')} vs {row.get('base_product', '')}"
            rows.append(
                {
                    "chart_id": "task2_product_tightening",
                    "task": "Task2",
                    "series": str(row.get("new_product", "")),
                    "category": label,
                    "metric": "tightened_count",
                    "value": int(row.get("tightened_count", 0)),
                    "label": label,
                    "color": colors[(index + 3) % len(colors)],
                    "sort_order": index + 1,
                    "include": True,
                    "title": "产品迭代指标收严成果",
                    "subtitle": "最后版本中后推出产品相较前序产品的可比规格收严项数",
                }
            )
        return pd.DataFrame(rows)

    def _write_csv_outputs(
        self,
        extracted_specs_df: pd.DataFrame,
        task1_summary_df: pd.DataFrame,
        task1_details_df: pd.DataFrame,
        task1_all_df: pd.DataFrame,
        task2_summary_df: pd.DataFrame,
        task2_details_df: pd.DataFrame,
        task2_all_df: pd.DataFrame,
        plot_config_df: pd.DataFrame,
    ) -> None:
        frames = {
            "extracted_specs.csv": extracted_specs_df,
            "task1_summary.csv": task1_summary_df,
            "task1_details.csv": task1_details_df,
            "task1_all_comparisons.csv": task1_all_df,
            "task2_summary.csv": task2_summary_df,
            "task2_details.csv": task2_details_df,
            "task2_all_comparisons.csv": task2_all_df,
            "plot_config.csv": plot_config_df,
        }
        try:
            for filename, frame in frames.items():
                frame.to_csv(self.output_dir / filename, index=False, encoding="utf-8-sig")
        except Exception as exc:
            logger.exception("[IndicatorImprovement] CSV 输出失败: %s", exc)
            raise

    def _write_excel_outputs(
        self,
        extracted_specs_df: pd.DataFrame,
        task1_summary_df: pd.DataFrame,
        task1_details_df: pd.DataFrame,
        task1_all_df: pd.DataFrame,
        task2_summary_df: pd.DataFrame,
        task2_details_df: pd.DataFrame,
        task2_all_df: pd.DataFrame,
        plot_config_df: pd.DataFrame,
        metadata_df: pd.DataFrame,
    ) -> Path:
        workbook_path = self.output_dir / "indicator_improvement_results.xlsx"
        try:
            with pd.ExcelWriter(workbook_path, engine="xlsxwriter") as writer:
                task1_summary_df.to_excel(writer, sheet_name="Task1_summary", index=False)
                task1_details_df.to_excel(writer, sheet_name="Task1_details", index=False)
                task1_all_df.to_excel(writer, sheet_name="Task1_all_compare", index=False)
                task2_summary_df.to_excel(writer, sheet_name="Task2_summary", index=False)
                task2_details_df.to_excel(writer, sheet_name="Task2_details", index=False)
                task2_all_df.to_excel(writer, sheet_name="Task2_all_compare", index=False)
                plot_config_df.to_excel(writer, sheet_name="Plot_config", index=False)
                extracted_specs_df.to_excel(writer, sheet_name="Extracted_specs", index=False)
                metadata_df.to_excel(writer, sheet_name="Run_metadata", index=False)
                _format_workbook(writer)
            return workbook_path
        except Exception as exc:
            logger.exception("[IndicatorImprovement] Excel 输出失败: %s", exc)
            raise

    def _draw_task1_chart(self, plot_config_df: pd.DataFrame) -> Path:
        image_path = self.output_dir / "task1_version_tightening.png"
        data = plot_config_df[
            (plot_config_df["chart_id"] == "task1_version_tightening") & (plot_config_df["include"] == True)
        ].copy()
        _draw_bar_chart(
            data=data,
            image_path=image_path,
            title="不同版本指标收严成果",
            subtitle="同一产品相较上一版本的可比规格收严项数",
        )
        return image_path

    def _draw_task2_chart(self, plot_config_df: pd.DataFrame) -> Path:
        image_path = self.output_dir / "task2_product_tightening.png"
        data = plot_config_df[
            (plot_config_df["chart_id"] == "task2_product_tightening") & (plot_config_df["include"] == True)
        ].copy()
        _draw_bar_chart(
            data=data,
            image_path=image_path,
            title="产品迭代指标收严成果",
            subtitle="最后版本中后推出产品相较前序产品的可比规格收严项数",
        )
        return image_path

    def _build_metadata(self, source_files: list[Path], extracted_specs_df: pd.DataFrame) -> pd.DataFrame:
        rows = [
            {"key": "source_dir", "value": str(self.source_dir)},
            {"key": "output_dir", "value": str(self.output_dir)},
            {"key": "source_files", "value": "\n".join(path.name for path in source_files)},
            {"key": "early_version_product_map", "value": json.dumps(self.early_version_product_map, ensure_ascii=False)},
            {"key": "products", "value": ",".join(self.products)},
            {"key": "extracted_spec_rows", "value": str(len(extracted_specs_df))},
            {
                "key": "comparison_rule",
                "value": "复杂或不可比规格不计入收严；可比较上下限/公差/范围时才判断收严。",
            },
        ]
        return pd.DataFrame(rows)

    def _write_metadata_json(
        self,
        source_files: list[Path],
        extracted_specs_df: pd.DataFrame,
        task1_summary_df: pd.DataFrame,
        task2_summary_df: pd.DataFrame,
    ) -> None:
        metadata = {
            "source_dir": str(self.source_dir),
            "output_dir": str(self.output_dir),
            "source_files": [path.name for path in source_files],
            "early_version_product_map": self.early_version_product_map,
            "products": list(self.products),
            "extracted_spec_rows": int(len(extracted_specs_df)),
            "task1_summary": task1_summary_df.to_dict(orient="records"),
            "task2_summary": task2_summary_df.to_dict(orient="records"),
        }
        try:
            (self.output_dir / "run_metadata.json").write_text(
                json.dumps(metadata, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as exc:
            logger.exception("[IndicatorImprovement] metadata 输出失败: %s", exc)
            raise


def _extract_version(filename: str) -> Optional[str]:
    match = re.search(r"(20\d{6})", filename)
    return match.group(1) if match else None


def _read_workbook_with_openpyxl(path: Path) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return {
                sheet_name: [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
                for sheet_name in workbook.sheetnames
            }
        finally:
            workbook.close()
    except (BadZipFile, OSError, ValueError) as exc:
        logger.warning("[IndicatorImprovement] openpyxl workbook error: %s", exc)
        raise
    except Exception as exc:
        logger.warning("[IndicatorImprovement] openpyxl unexpected error: %s", exc)
        raise


def _read_workbook_with_com(path: Path) -> dict[str, list[list[Any]]]:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Excel COM 读取需要 pywin32 和本地 Excel。") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()))
        sheets: dict[str, list[list[Any]]] = {}
        for index in range(1, workbook.Sheets.Count + 1):
            worksheet = workbook.Sheets(index)
            used_range = worksheet.UsedRange
            raw_data = used_range.Value
            sheets[str(worksheet.Name)] = _com_range_to_rows(raw_data)
        return sheets
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _save_workbook_with_com(source_path: Path, output_path: Path) -> None:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Excel COM 保存需要 pywin32 和本地 Excel。") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source_path.resolve()))
        workbook.SaveAs(str(output_path.resolve()), FileFormat=51)
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _com_range_to_rows(raw_data: Any) -> list[list[Any]]:
    if raw_data is None:
        return []
    if not isinstance(raw_data, tuple):
        return [[raw_data]]
    if raw_data and not isinstance(raw_data[0], tuple):
        return [list(raw_data)]
    return [list(row) for row in raw_data]


def _find_header_index(rows: list[list[Any]]) -> Optional[int]:
    for index, row in enumerate(rows[:20]):
        values = {normalize_cell(value) for value in row}
        if all(required in values for required in REQUIRED_COLUMNS):
            return index
    return None


def _index_records(records: list[IndicatorSpecRecord]) -> dict[tuple[str, str], dict[str, IndicatorSpecRecord]]:
    indexed: dict[tuple[str, str], dict[str, IndicatorSpecRecord]] = {}
    for record in records:
        key = (record.product, record.version)
        indexed.setdefault(key, {})[record.identity_key] = record
    return indexed


def _compare_record_maps(
    task: str,
    comparison_label: str,
    old_label: str,
    new_label: str,
    old_records: dict[str, IndicatorSpecRecord],
    new_records: dict[str, IndicatorSpecRecord],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    common_keys = sorted(set(old_records) & set(new_records))
    for identity_key in common_keys:
        old_record = old_records[identity_key]
        new_record = new_records[identity_key]
        comparison = compare_specs(old_record.raw_spec, new_record.raw_spec)
        rows.append(
            {
                "task": task,
                "comparison": comparison_label,
                "product": new_record.product,
                "from_version": old_label if old_label.isdigit() else "",
                "to_version": new_label if new_label.isdigit() else "",
                "from_label": old_label,
                "to_label": new_label,
                "identity_key": identity_key,
                "spec_name": new_record.display_name,
                "factory": new_record.factory,
                "department": new_record.department,
                "process_layer": new_record.process_layer,
                "station": new_record.station,
                "monitor_factor": new_record.monitor_factor,
                "rs_code": new_record.rs_code,
                "description": new_record.description,
                "old_spec": old_record.raw_spec,
                "new_spec": new_record.raw_spec,
                "old_parsed": constraints_to_text(comparison.old_spec.constraints),
                "new_parsed": constraints_to_text(comparison.new_spec.constraints),
                "old_comparable": comparison.old_spec.comparable,
                "new_comparable": comparison.new_spec.comparable,
                "is_tightened": comparison.is_tightened,
                "reason": comparison.reason,
            }
        )
    return rows


def _summarize_comparisons(all_df: pd.DataFrame, group_columns: list[str]) -> pd.DataFrame:
    if all_df.empty:
        columns = group_columns + ["comparison", "tightened_count", "comparable_count", "compared_count", "tightened_names"]
        return pd.DataFrame(columns=columns)

    rows: list[dict[str, Any]] = []
    for keys, group in all_df.groupby(group_columns, dropna=False):
        if not isinstance(keys, tuple):
            keys = (keys,)
        tightened = group[group["is_tightened"] == True]
        comparable = group[(group["old_comparable"] == True) & (group["new_comparable"] == True)]
        row = {column: key for column, key in zip(group_columns, keys)}
        row["comparison"] = str(group["comparison"].iloc[0])
        row["tightened_count"] = int(len(tightened))
        row["comparable_count"] = int(len(comparable))
        row["compared_count"] = int(len(group))
        row["tightened_names"] = "\n".join(tightened["spec_name"].astype(str).tolist())
        rows.append(row)
    return pd.DataFrame(rows).sort_values(group_columns).reset_index(drop=True)


def _format_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_format = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#1F4E78", "border": 1, "align": "center"}
    )
    wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})
    for worksheet in writer.sheets.values():
        worksheet.freeze_panes(1, 0)
        worksheet.set_row(0, None, header_format)
        worksheet.set_column(0, 2, 18, wrap_format)
        worksheet.set_column(3, 12, 24, wrap_format)
        worksheet.set_column(13, 30, 34, wrap_format)


def _draw_bar_chart(data: pd.DataFrame, image_path: Path, title: str, subtitle: str) -> None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
    except Exception as exc:
        logger.exception("[IndicatorImprovement] 绘图依赖加载失败: %s", exc)
        raise

    font_name = _pick_chinese_font(font_manager)
    if font_name:
        plt.rcParams["font.sans-serif"] = [font_name]
    plt.rcParams["axes.unicode_minus"] = False

    data = data.sort_values("sort_order")
    categories = data["label"].astype(str).tolist()
    values = data["value"].astype(float).tolist()
    colors = data["color"].astype(str).tolist()

    fig, ax = plt.subplots(figsize=(13.5, 7.5), dpi=180)
    fig.patch.set_facecolor("#F6F8FB")
    ax.set_facecolor("#F6F8FB")
    bars = ax.barh(categories, values, color=colors, height=0.56)
    ax.invert_yaxis()
    ax.grid(axis="x", color="#D8DEE9", linewidth=0.8, alpha=0.8)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.tick_params(axis="both", length=0, labelsize=11, colors="#263238")
    ax.set_xlabel("收严规格项数", fontsize=12, color="#37474F", labelpad=12)
    max_value = max(values) if values else 0
    ax.set_xlim(0, max(1, max_value * 1.18))
    for bar, value in zip(bars, values):
        ax.text(
            value + max(0.2, max_value * 0.02),
            bar.get_y() + bar.get_height() / 2,
            f"{int(value)}",
            va="center",
            ha="left",
            fontsize=13,
            fontweight="bold",
            color="#102027",
        )
    fig.text(0.06, 0.94, title, fontsize=24, fontweight="bold", color="#102027")
    fig.text(0.06, 0.89, subtitle, fontsize=12, color="#546E7A")
    fig.text(0.06, 0.055, "数据来源：维信诺北极星PNL指标规格表；复杂不可比规格按规则不计入收严。", fontsize=9, color="#78909C")
    plt.tight_layout(rect=(0.055, 0.08, 0.98, 0.86))
    try:
        fig.savefig(image_path, facecolor=fig.get_facecolor(), bbox_inches="tight")
    finally:
        plt.close(fig)


def _pick_chinese_font(font_manager: Any) -> Optional[str]:
    preferred = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Source Han Sans SC", "Arial Unicode MS"]
    available = {font.name for font in font_manager.fontManager.ttflist}
    for name in preferred:
        if name in available:
            return name
    return None

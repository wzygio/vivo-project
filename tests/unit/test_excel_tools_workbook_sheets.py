import os
from pathlib import Path

import pandas as pd
import pytest

from src.shared_kernel.utils.excel_tools import (
    WorkbookWriteResult,
    list_workbook_sheet_names,
    read_workbook_sheet,
    replace_workbook_sheet,
    replace_workbook_sheets,
)


def test_read_workbook_sheet_returns_empty_for_missing_file(tmp_path: Path) -> None:
    result = read_workbook_sheet(tmp_path / "missing.xlsx", "M678")
    assert result.empty


def test_read_workbook_sheet_returns_empty_for_missing_sheet(tmp_path: Path) -> None:
    workbook = tmp_path / "shared.xlsx"
    pd.DataFrame({"a": [1]}).to_excel(workbook, index=False, sheet_name="M626")

    result = read_workbook_sheet(workbook, "M678")
    assert result.empty


def test_read_workbook_sheet_reads_only_target_sheet(tmp_path: Path) -> None:
    workbook = tmp_path / "shared.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"a": [1]}).to_excel(writer, index=False, sheet_name="M626")
        pd.DataFrame({"b": [2, 3]}).to_excel(writer, index=False, sheet_name="M678")

    result = read_workbook_sheet(workbook, "M678")
    pd.testing.assert_frame_equal(result, pd.DataFrame({"b": [2, 3]}))


def test_read_workbook_sheet_falls_back_to_com_for_encrypted_file(monkeypatch, tmp_path: Path) -> None:
    workbook = tmp_path / "encrypted.xlsx"
    workbook.write_bytes(b"\x00not-a-real-xlsx")
    expected = pd.DataFrame({"a": [1]})

    monkeypatch.setattr(pd, "read_excel", lambda *a, **k: (_ for _ in ()).throw(RuntimeError("encrypted")))
    import src.shared_kernel.utils.excel_tools as excel_tools

    monkeypatch.setattr(excel_tools, "_read_encrypted_xlsx_via_com", lambda path, sheet_name=None: expected)

    result = read_workbook_sheet(workbook, "M678")
    pd.testing.assert_frame_equal(result, expected)


def test_replace_workbook_sheet_creates_new_workbook(tmp_path: Path) -> None:
    workbook = tmp_path / "shared.xlsx"
    replace_workbook_sheet(workbook, "M678", pd.DataFrame({"a": [1, 2]}))

    assert workbook.exists()
    result = pd.read_excel(workbook, sheet_name="M678")
    pd.testing.assert_frame_equal(result, pd.DataFrame({"a": [1, 2]}))


def test_replace_workbook_sheet_preserves_other_sheets(tmp_path: Path) -> None:
    workbook = tmp_path / "shared.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"keep": [1]}).to_excel(writer, index=False, sheet_name="M626")
        pd.DataFrame({"old": [9]}).to_excel(writer, index=False, sheet_name="M678")

    replace_workbook_sheet(workbook, "M678", pd.DataFrame({"new": [5]}))

    preserved = pd.read_excel(workbook, sheet_name="M626")
    replaced = pd.read_excel(workbook, sheet_name="M678")
    pd.testing.assert_frame_equal(preserved, pd.DataFrame({"keep": [1]}))
    pd.testing.assert_frame_equal(replaced, pd.DataFrame({"new": [5]}))


def test_replace_workbook_sheet_skips_locked_workbook_without_raising(monkeypatch, tmp_path: Path) -> None:
    workbook = tmp_path / "shared.xlsx"
    original = pd.DataFrame({"a": [1]})
    original.to_excel(workbook, index=False, sheet_name="M678")

    import openpyxl

    def _locked_save(self, path):
        raise PermissionError("locked")

    monkeypatch.setattr(openpyxl.Workbook, "save", _locked_save)

    replace_workbook_sheet(workbook, "M678", pd.DataFrame({"a": [2]}))

    # 原内容未被破坏
    pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M678"), original)


def test_replace_workbook_sheet_rewrites_encrypted_workbook_via_com(monkeypatch, tmp_path: Path) -> None:
    workbook = tmp_path / "encrypted.xlsx"
    workbook.write_bytes(b"\x00enterprise-encrypted")

    import openpyxl
    import src.shared_kernel.utils.excel_tools as excel_tools

    # 仅对原文件模拟“企业加密打不开”，临时文件仍需可被 openpyxl 正常验证
    real_load_workbook = openpyxl.load_workbook

    def _load_workbook(path, *args, **kwargs):
        if Path(path) == workbook:
            raise RuntimeError("encrypted")
        return real_load_workbook(path, *args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", _load_workbook)
    monkeypatch.setattr(
        excel_tools,
        "_read_all_sheets_via_com",
        lambda path: {"M626": pd.DataFrame({"keep": [1]}), "M678": pd.DataFrame({"old": [9]})},
    )

    replace_workbook_sheet(workbook, "M678", pd.DataFrame({"new": [5]}))

    monkeypatch.undo()  # 恢复 openpyxl.load_workbook，以便真实读回结果

    pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M626"), pd.DataFrame({"keep": [1]}))
    pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M678"), pd.DataFrame({"new": [5]}))


class TestReplaceWorkbookSheets:
    """多 sheet 原子写入与 WorkbookWriteResult 契约测试（PRD §5.10 / 验收 11.4）。"""

    def test_updates_multiple_sheets_and_preserves_others(self, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
            pd.DataFrame({"keep": [1]}).to_excel(writer, index=False, sheet_name="M626")
            pd.DataFrame({"old": [9]}).to_excel(writer, index=False, sheet_name="M678")

        result = replace_workbook_sheets(
            workbook,
            {
                "M678": pd.DataFrame({"new": [5]}),
                "Z517": pd.DataFrame({"z": [7, 8]}),
            },
        )

        assert result.written is True
        assert result.error is None
        assert result.path == workbook
        assert result.updated_sheets == ("M678", "Z517")
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M626"), pd.DataFrame({"keep": [1]}))
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M678"), pd.DataFrame({"new": [5]}))
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="Z517"), pd.DataFrame({"z": [7, 8]}))

    def test_creates_workbook_when_missing(self, tmp_path: Path) -> None:
        workbook = tmp_path / "new.xlsx"

        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"a": [1, 2]})})

        assert result.written is True
        assert result.updated_sheets == ("M678",)
        assert workbook.exists()
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M678"), pd.DataFrame({"a": [1, 2]}))

    def test_permission_error_on_replace_returns_written_false(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        original = pd.DataFrame({"a": [1]})
        original.to_excel(workbook, index=False, sheet_name="M678")
        original_bytes = workbook.read_bytes()

        def _locked_replace(src, dst):
            raise PermissionError("locked by Excel")

        monkeypatch.setattr(os, "replace", _locked_replace)

        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"a": [2]})})

        assert result.written is False
        assert result.error is not None
        assert "请关闭 Excel 后重试" in result.error
        # 正式文件字节不变，且不留临时文件
        assert workbook.read_bytes() == original_bytes
        assert not list(tmp_path.glob("*.tmp.xlsx"))

    def test_temp_save_failure_keeps_original_bytes(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        original = pd.DataFrame({"a": [1]})
        original.to_excel(workbook, index=False, sheet_name="M678")
        original_bytes = workbook.read_bytes()

        import openpyxl

        def _broken_save(self, path):
            raise RuntimeError("disk full")

        monkeypatch.setattr(openpyxl.Workbook, "save", _broken_save)

        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"a": [2]})})

        assert result.written is False
        assert result.error is not None
        assert workbook.read_bytes() == original_bytes
        assert not list(tmp_path.glob("*.tmp.xlsx"))

    def test_verification_failure_does_not_replace(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        original = pd.DataFrame({"a": [1]})
        original.to_excel(workbook, index=False, sheet_name="M678")
        original_bytes = workbook.read_bytes()

        import openpyxl

        # 临时文件回读验证时模拟“目标 sheet 缺失”的损坏结果
        real_load_workbook = openpyxl.load_workbook
        calls = {"n": 0}

        def _load_workbook(path, *args, **kwargs):
            calls["n"] += 1
            wb = real_load_workbook(path, *args, **kwargs)
            if calls["n"] > 1 and "M678" in wb.sheetnames:
                del wb["M678"]
            return wb

        monkeypatch.setattr(openpyxl, "load_workbook", _load_workbook)

        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"a": [2]})})

        assert result.written is False
        assert result.error is not None
        assert workbook.read_bytes() == original_bytes
        assert not list(tmp_path.glob("*.tmp.xlsx"))

    def test_encrypted_fallback_via_com_succeeds(self, monkeypatch, tmp_path: Path, caplog) -> None:
        workbook = tmp_path / "encrypted.xlsx"
        workbook.write_bytes(b"\x00enterprise-encrypted")

        import openpyxl
        import src.shared_kernel.utils.excel_tools as excel_tools

        # 仅对原文件模拟“企业加密打不开”，临时文件仍需可被 openpyxl 正常验证
        real_load_workbook = openpyxl.load_workbook

        def _load_workbook(path, *args, **kwargs):
            if Path(path) == workbook:
                raise RuntimeError("encrypted")
            return real_load_workbook(path, *args, **kwargs)

        monkeypatch.setattr(openpyxl, "load_workbook", _load_workbook)
        monkeypatch.setattr(
            excel_tools,
            "_read_all_sheets_via_com",
            lambda path: {"M626": pd.DataFrame({"keep": [1]}), "M678": pd.DataFrame({"old": [9]})},
        )

        with caplog.at_level("WARNING"):
            result = replace_workbook_sheets(
                workbook,
                {"M678": pd.DataFrame({"new": [5]}), "Z517": pd.DataFrame({"z": [7]})},
            )

        assert result.written is True
        assert result.updated_sheets == ("M678", "Z517")
        # 加密工作簿整体重写为明文必须有明确告警
        assert any("明文" in record.message for record in caplog.records)

        monkeypatch.undo()  # 恢复 openpyxl.load_workbook，以便真实读回结果

        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M626"), pd.DataFrame({"keep": [1]}))
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="M678"), pd.DataFrame({"new": [5]}))
        pd.testing.assert_frame_equal(pd.read_excel(workbook, sheet_name="Z517"), pd.DataFrame({"z": [7]}))

    def test_com_read_failure_returns_written_false(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "encrypted.xlsx"
        original_bytes = b"\x00enterprise-encrypted"
        workbook.write_bytes(original_bytes)

        import openpyxl
        import src.shared_kernel.utils.excel_tools as excel_tools

        monkeypatch.setattr(
            openpyxl, "load_workbook", lambda *a, **k: (_ for _ in ()).throw(RuntimeError("encrypted"))
        )

        def _broken_com(path):
            raise RuntimeError("COM unavailable")

        monkeypatch.setattr(excel_tools, "_read_all_sheets_via_com", _broken_com)

        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"new": [5]})})

        assert result.written is False
        assert result.error is not None
        # 正式文件字节不变，且不留临时文件
        assert workbook.read_bytes() == original_bytes
        assert not list(tmp_path.glob("*.tmp.xlsx"))

    def test_result_is_frozen_dataclass(self, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        result = replace_workbook_sheets(workbook, {"M678": pd.DataFrame({"a": [1]})})

        assert isinstance(result, WorkbookWriteResult)
        with pytest.raises(AttributeError):
            result.written = False


class TestListWorkbookSheetNames:
    """list_workbook_sheet_names：openpyxl 直通 / COM 回退 / 双失败三分支。"""

    def test_returns_empty_list_for_missing_file(self, tmp_path: Path) -> None:
        assert list_workbook_sheet_names(tmp_path / "missing.xlsx") == []

    def test_lists_sheet_names_via_openpyxl(self, tmp_path: Path) -> None:
        workbook = tmp_path / "shared.xlsx"
        with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
            pd.DataFrame({"a": [1]}).to_excel(writer, index=False, sheet_name="M626")
            pd.DataFrame({"b": [2]}).to_excel(writer, index=False, sheet_name="M678__flags")

        assert list_workbook_sheet_names(workbook) == ["M626", "M678__flags"]

    def test_falls_back_to_com_for_encrypted_file(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "encrypted.xlsx"
        workbook.write_bytes(b"\x00enterprise-encrypted")

        import openpyxl
        import src.shared_kernel.utils.excel_tools as excel_tools

        monkeypatch.setattr(
            openpyxl, "load_workbook",
            lambda *a, **k: (_ for _ in ()).throw(RuntimeError("encrypted")),
        )
        monkeypatch.setattr(
            excel_tools, "_list_sheet_names_via_com",
            lambda path: ["M678", "M678__flags"],
        )

        assert list_workbook_sheet_names(workbook) == ["M678", "M678__flags"]

    def test_returns_none_when_openpyxl_and_com_both_fail(self, monkeypatch, tmp_path: Path) -> None:
        workbook = tmp_path / "encrypted.xlsx"
        workbook.write_bytes(b"\x00enterprise-encrypted")

        import openpyxl
        import src.shared_kernel.utils.excel_tools as excel_tools

        monkeypatch.setattr(
            openpyxl, "load_workbook",
            lambda *a, **k: (_ for _ in ()).throw(RuntimeError("encrypted")),
        )

        def _broken_com(path):
            raise RuntimeError("COM unavailable")

        monkeypatch.setattr(excel_tools, "_list_sheet_names_via_com", _broken_com)

        assert list_workbook_sheet_names(workbook) is None


def test_com_read_returns_empty_for_missing_sheet(monkeypatch, tmp_path: Path) -> None:
    """COM 读取指定 sheet 缺失时返回空 DataFrame（与明文读取语义一致）。"""
    import sys
    import types

    import src.shared_kernel.utils.excel_tools as excel_tools

    workbook = tmp_path / "encrypted.xlsx"
    workbook.write_bytes(b"\x00enterprise-encrypted")

    class _FakeWorksheet:
        def __init__(self, name):
            self.Name = name

    class _FakeWorksheets:
        def __init__(self, names):
            self._sheets = [_FakeWorksheet(name) for name in names]

        def __iter__(self):
            return iter(self._sheets)

        def __call__(self, key):
            for ws in self._sheets:
                if ws.Name == key:
                    return ws
            raise KeyError(key)

    class _FakeWorkbook:
        def __init__(self, names):
            self.Worksheets = _FakeWorksheets(names)

        def Close(self, SaveChanges=False):
            pass

    class _FakeWorkbooks:
        def __init__(self, wb):
            self._wb = wb

        def Open(self, path, ReadOnly=True):
            return self._wb

    class _FakeExcel:
        def __init__(self, wb):
            self.Workbooks = _FakeWorkbooks(wb)
            self.Visible = False
            self.DisplayAlerts = False

        def Quit(self):
            pass

    fake_wb = _FakeWorkbook(["M626"])
    fake_client = types.ModuleType("win32com.client")
    fake_client.DispatchEx = lambda app: _FakeExcel(fake_wb)
    fake_win32com = types.ModuleType("win32com")
    fake_win32com.client = fake_client
    fake_pythoncom = types.ModuleType("pythoncom")
    fake_pythoncom.CoInitialize = lambda: None
    fake_pythoncom.CoUninitialize = lambda: None

    monkeypatch.setitem(sys.modules, "win32com", fake_win32com)
    monkeypatch.setitem(sys.modules, "win32com.client", fake_client)
    monkeypatch.setitem(sys.modules, "pythoncom", fake_pythoncom)

    result = excel_tools._read_encrypted_xlsx_via_com(workbook, "M678")
    assert result.empty

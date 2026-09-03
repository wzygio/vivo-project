"""企业加密工作簿场景：sheet 名枚举 COM 回退后的门控/首写/读取语义。

企业加密的模拟方式（不依赖真实 Excel）：monkeypatch openpyxl.load_workbook
使其对目标工作簿抛错，excel_tools 的 COM 枚举/读取回退改用未打补丁的真实
openpyxl 句柄实现——等价于"COM 能透明解密读取"的行为契约。

覆盖（缺陷 1/2 修复回归）：
- 加密 + __flags 已存在 → 4h 门控生效，persist 不重写、不覆写 __flags；
- 加密 + __flags 缺失但旧产品 sheet 存在 → 首写物化空 __flags（不继承旧表 flag），幂等；
- 加密 + __flags 缺失 → load_sheet_oos_decisions 返回空台账而非抛错；
- openpyxl 与 COM 都无法枚举 → 读/persist 一律抛 SheetOosDecorationReadError。
"""

from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest

from src.inline_domain.core.shared.sheet_oos_decoration import (
    EMPTY_DECISION_SIGNATURE,
    OOS_DECORATION_FILE_NAME,
    OOS_KEY_COLUMNS,
    REFRESH_META_SHEET_NAME,
    build_refresh_meta_row,
    build_sheet_oos_detail,
    compute_decision_signature,
)
from src.inline_domain.infrastructure.shared import (
    sheet_oos_decoration_repository as sheet_oos_decoration,
)
from src.inline_domain.infrastructure.shared.sheet_oos_decoration_repository import (
    SheetOosDecorationReadError,
    load_sheet_oos_decisions,
    persist_sheet_oos_decoration,
)
from src.shared_kernel.utils.excel_tools import (
    list_workbook_sheet_names,
    replace_workbook_sheets,
)

NOW = datetime(2026, 8, 18, 12, 0, 0)
SHEET = "Z571"
SCOPE = "spc"


def _sheet_features() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "factory": "OLED",
                "prod_code": "Z571",
                "step_id": "21200",
                "param_name": "PPA_B_X",
                "sheet_id": "S1",
                "sheet_start_time": "2026-07-01 08:00:00",
                "sheet_max": 7.5,
                "sheet_min": -1.0,
                "sheet_mean": 0.5,
                "usl": 6.0,
                "lsl": -6.0,
            },
            {
                "factory": "OLED",
                "prod_code": "Z571",
                "step_id": "21200",
                "param_name": "PPA_B_X",
                "sheet_id": "S2",
                "sheet_start_time": "2026-07-01 09:00:00",
                "sheet_max": 2.0,
                "sheet_min": -7.5,
                "sheet_mean": -0.5,
                "usl": 6.0,
                "lsl": -6.0,
            },
        ]
    )


def _decisions_ledger() -> pd.DataFrame:
    """用户维护的决策台账：S1 被管理员关闭修饰（False），S2 保持 True。"""
    return pd.DataFrame(
        [
            {"prod_code": "Z571", "step_id": "21200", "param_name": "PPA_B_X", "sheet_id": "S1", "flag": False},
            {"prod_code": "Z571", "step_id": "21200", "param_name": "PPA_B_X", "sheet_id": "S2", "flag": True},
        ]
    )


def _legacy_product_sheet() -> pd.DataFrame:
    """旧格式产品 sheet（明细列 + flag）：S1 flag=False，S2 flag=True。"""
    return pd.DataFrame(
        [
            {"prod_code": "Z571", "step_id": "21200", "param_name": "PPA_B_X",
             "sheet_id": "S1", "sheet_max": 7.5, "flag": False},
            {"prod_code": "Z571", "step_id": "21200", "param_name": "PPA_B_X",
             "sheet_id": "S2", "sheet_max": 2.0, "flag": True},
        ]
    )


def _make_product_dir(tmp_path: Path) -> Path:
    product_dir = tmp_path / "resources"
    product_dir.mkdir(parents=True, exist_ok=True)
    return product_dir


def _write_workbook(product_dir: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    workbook = product_dir / OOS_DECORATION_FILE_NAME
    result = replace_workbook_sheets(workbook, sheets)
    assert result.written, result.error
    return workbook


def _spy_writes(monkeypatch) -> list[tuple[str, ...]]:
    """拦截模块内的 replace_workbook_sheets，记录每次写入的 sheet 名并继续真实写入。"""
    write_calls: list[tuple[str, ...]] = []
    real_replace = sheet_oos_decoration.replace_workbook_sheets

    def spy(path, sheets):
        write_calls.append(tuple(sheets.keys()))
        return real_replace(path, sheets)

    monkeypatch.setattr(sheet_oos_decoration, "replace_workbook_sheets", spy)
    return write_calls


def _simulate_encrypted_workbook(monkeypatch, workbook: Path):
    """模拟企业加密：openpyxl 打不开目标文件；COM 回退可透明读取真实内容。

    返回 COM 单 sheet 读取 mock，供测试在读侧断言加密工作簿内的真实状态。
    """
    import openpyxl
    import src.shared_kernel.utils.excel_tools as excel_tools

    real_load_workbook = openpyxl.load_workbook

    def _load_workbook(path, *args, **kwargs):
        if Path(path) == Path(workbook):
            raise RuntimeError("enterprise-encrypted")
        return real_load_workbook(path, *args, **kwargs)

    def _com_sheet_names(path: Path) -> list[str]:
        wb = real_load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def _com_read_sheet(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
        wb = real_load_workbook(path, read_only=True)
        try:
            target = sheet_name or wb.sheetnames[0]
            if target not in wb.sheetnames:
                return pd.DataFrame()
            rows = list(wb[target].values)
        finally:
            wb.close()
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows[1:], columns=list(rows[0]))

    def _com_read_all(path: Path) -> dict[str, pd.DataFrame]:
        return {name: _com_read_sheet(path, name) for name in _com_sheet_names(path)}

    monkeypatch.setattr(openpyxl, "load_workbook", _load_workbook)
    monkeypatch.setattr(excel_tools, "_list_sheet_names_via_com", _com_sheet_names)
    monkeypatch.setattr(excel_tools, "_read_encrypted_xlsx_via_com", _com_read_sheet)
    monkeypatch.setattr(excel_tools, "_read_all_sheets_via_com", _com_read_all)
    # sheet_oos_decoration 以模块级名字直接引用 COM 读取函数，需同步替换
    monkeypatch.setattr(sheet_oos_decoration, "_read_encrypted_xlsx_via_com", _com_read_sheet)
    return _com_read_sheet


def _persist_kwargs(**overrides) -> dict:
    kwargs = dict(
        sheet_name=SHEET,
        scope=SCOPE,
        product_revision="R1",
        decision_signature="sig-1",
        now=NOW,
    )
    kwargs.update(overrides)
    return kwargs


# ---------------------------------------------------------------------------
# a) 加密 + __flags 已存在：门控生效，不重写、不覆写 __flags
# ---------------------------------------------------------------------------


def test_encrypted_workbook_with_flags_sheet_skips_rewrite_within_ttl(
    tmp_path: Path, monkeypatch
) -> None:
    product_dir = _make_product_dir(tmp_path)
    detail = build_sheet_oos_detail(_sheet_features())
    meta_row = build_refresh_meta_row(
        scope=SCOPE,
        prod_code=SHEET,
        generated_at=NOW,
        product_revision="R1",
        decision_signature="sig-1",
        detail_row_count=2,
    )
    workbook = _write_workbook(
        product_dir,
        {
            SHEET: detail.assign(flag=[False, True]),
            f"{SHEET}__flags": _decisions_ledger(),
            REFRESH_META_SHEET_NAME: pd.DataFrame([meta_row]),
        },
    )
    com_read_sheet = _simulate_encrypted_workbook(monkeypatch, workbook)
    write_calls = _spy_writes(monkeypatch)
    bytes_before = workbook.read_bytes()

    merged = persist_sheet_oos_decoration(
        product_dir, detail, **_persist_kwargs(now=NOW + timedelta(hours=1))
    )

    # 4h 门控对加密文件同样生效：TTL 内不重写，文件字节不变
    assert write_calls == []
    assert workbook.read_bytes() == bytes_before
    # 决策来自用户维护的 __flags（S1=False），而非系统默认 True
    assert merged["flag"].tolist() == [False, True]

    # TTL 到期 → 只重写产品 sheet + meta；__flags 不在写入集合，用户决策不被覆写
    persist_sheet_oos_decoration(
        product_dir, detail, **_persist_kwargs(now=NOW + timedelta(hours=4))
    )
    assert write_calls == [(SHEET, REFRESH_META_SHEET_NAME)]
    flags_after = com_read_sheet(workbook, f"{SHEET}__flags")
    assert flags_after["flag"].tolist() == [False, True]


# ---------------------------------------------------------------------------
# b) 加密 + __flags 缺失但旧产品 sheet 存在：首写生成空 __flags（不继承），幂等
# ---------------------------------------------------------------------------


def test_encrypted_workbook_first_write_creates_empty_flags_idempotent(
    tmp_path: Path, monkeypatch
) -> None:
    """旧产品 sheet 的 flag 永不生效：首写物化空 __flags，全部键默认 True。"""
    product_dir = _make_product_dir(tmp_path)
    detail = build_sheet_oos_detail(_sheet_features())
    workbook = _write_workbook(product_dir, {SHEET: _legacy_product_sheet()})
    com_read_sheet = _simulate_encrypted_workbook(monkeypatch, workbook)
    write_calls = _spy_writes(monkeypatch)

    merged = persist_sheet_oos_decoration(product_dir, detail, **_persist_kwargs())

    # 首次运行：不继承旧产品 sheet 的 flag（S1=False 不生效），物化空 __flags
    assert write_calls == [(SHEET, f"{SHEET}__flags", REFRESH_META_SHEET_NAME)]
    assert merged["flag"].tolist() == [True, True]
    decisions = load_sheet_oos_decisions(product_dir, sheet_name=SHEET)
    assert decisions.empty
    assert list(decisions.columns) == [*OOS_KEY_COLUMNS, "flag"]

    # 二次运行：__flags 已存在，TTL 内不重写；旧表 flag 仍不生效
    bytes_before = workbook.read_bytes()
    merged_again = persist_sheet_oos_decoration(
        product_dir, detail, **_persist_kwargs(now=NOW + timedelta(hours=1))
    )
    assert len(write_calls) == 1
    assert workbook.read_bytes() == bytes_before
    assert merged_again["flag"].tolist() == [True, True]
    flags_after = com_read_sheet(workbook, f"{SHEET}__flags")
    assert flags_after.empty


# ---------------------------------------------------------------------------
# c) 加密 + __flags 缺失：load_sheet_oos_decisions 返回空台账而非抛错
# ---------------------------------------------------------------------------


def test_encrypted_workbook_without_flags_sheet_returns_empty_ledger(
    tmp_path: Path, monkeypatch
) -> None:
    product_dir = _make_product_dir(tmp_path)
    workbook = _write_workbook(product_dir, {SHEET: _legacy_product_sheet()})
    _simulate_encrypted_workbook(monkeypatch, workbook)

    decisions = load_sheet_oos_decisions(product_dir, sheet_name=SHEET)

    assert decisions.empty
    assert list(decisions.columns) == [*OOS_KEY_COLUMNS, "flag"]
    # 决策签名走确定性空值，全新加密工作簿不再向页面抛错
    assert compute_decision_signature(decisions) == EMPTY_DECISION_SIGNATURE


# ---------------------------------------------------------------------------
# d) openpyxl 与 COM 都无法枚举：显式抛 SheetOosDecorationReadError，不降级
# ---------------------------------------------------------------------------


def test_unreadable_workbook_raises_read_error(tmp_path: Path, monkeypatch) -> None:
    product_dir = _make_product_dir(tmp_path)
    detail = build_sheet_oos_detail(_sheet_features())
    workbook = _write_workbook(product_dir, {f"{SHEET}__flags": _decisions_ledger()})
    _simulate_encrypted_workbook(monkeypatch, workbook)

    import src.shared_kernel.utils.excel_tools as excel_tools

    def _broken_com_enum(path: Path) -> list[str]:
        raise RuntimeError("COM unavailable")

    monkeypatch.setattr(excel_tools, "_list_sheet_names_via_com", _broken_com_enum)
    bytes_before = workbook.read_bytes()

    # 双失败 → 枚举结果为 None（文件不可读），不得当作空清单
    assert list_workbook_sheet_names(workbook) is None

    with pytest.raises(SheetOosDecorationReadError):
        load_sheet_oos_decisions(product_dir, sheet_name=SHEET)
    # persist 路径不得静默继续（避免覆写不可读文件中的 __flags）
    with pytest.raises(SheetOosDecorationReadError):
        persist_sheet_oos_decoration(product_dir, detail, **_persist_kwargs())

    assert workbook.read_bytes() == bytes_before
